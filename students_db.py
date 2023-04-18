import sqlite3

import os
import secrets
import string

#import xlwings as xw
# note: xlwings does not work in ubuntu
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class StudentsDB:

    def __init__(self, db_file="students.db"):

        new_db = True

        if os.path.exists(db_file):
            new_db = False

        self.connection = sqlite3.connect(db_file)

        if new_db:
            create_table = """ CREATE TABLE students (
                            id INT NOT NULL UNIQUE,
                            username  VARCHAR(255) NOT NULL,
                            password  VARCHAR(255) NOT NULL,
                            user_group     VARCHAR(255) NOT NULL,
                            user_subgroup  VARCHAR(255) NOT NULL,
                            firstname      VARCHAR(255),
                            surname        VARCHAR(255),
                            matricola      VARCHAR(255),
                            docente        VARCHAR(255),
                            email          VARCHAR(255),
                            repository_url VARCHAR(255),
                            activated      TIMESTAMP
                        ); """

            self.connection.execute(create_table)

    def __del__(self):
        self.connection.close()

    def __iter__(self):
        return StudentsIter(self)

    def get_num_students(self) -> int:

        count_students = "SELECT COUNT(*) FROM students"

        num_students = self.connection.execute(count_students)

        return num_students.fetchone()[0]


    def set_repository_url(self, username: str, repository_url: str) -> None:

        set_url = "UPDATE students SET repository_url=? WHERE username=?"

        cursor = self.connection.cursor()
        cursor.execute(set_url, (repository_url,username))
        self.connection.commit()
        cursor.close()


    def initialize_users(self, num_students: int, prefix_username: str, password_length: int, top_project_group: str, project_subgroup: str, initial_user_id: int) -> None:

        alphabet = string.ascii_letters + string.digits

        count_students = "SELECT MAX(id) FROM students"
        max_students = self.connection.execute(count_students)
        student_id = max_students.fetchone() if max_students.rowcount > 0 else 1

        student_id = initial_user_id if initial_user_id > student_id else student_id

        cursor = self.connection.cursor()

        for i in range(1, num_students+1):

            username = prefix_username + str(student_id)

            password = ''.join(secrets.choice(alphabet) for i in range(password_length))

            print(f"Initializing user '{username}'")

            insert_student = "INSERT INTO students VALUES (?,?,?,?,?,'','','','','','',NULL)"

            cursor.execute(insert_student, (student_id, username, password, top_project_group, project_subgroup))

            student_id += 1

        self.connection.commit()
        cursor.close()


    def import_users(self, xlsx_path: str, prefix_username: str, password_length: int, top_project_group: str, project_subgroup: str, initial_user_id: int) -> None:

        if not os.path.exists(xlsx_path):
            raise ValueError(f"Error: '{xlsx_path}' does not exist")

        if not xlsx_path.endswith('.xlsx'):
            raise ValueError(f"Error: '{xlsx_path}' does not seem an XLSX file")


        #self.wb = xw.Book(xlsx_path)
        #self.sheet = self.wb.sheets[0]
        self.wb = load_workbook(xlsx_path)
        self.sheet = self.wb.worksheets[0]
        self.xlsx_path = xlsx_path

        #empty_column = self.sheet.used_range[-1].offset(column_offset=1).column
        empty_column = len(list(self.sheet.columns)) + 1

        self.username_column = None
        self.password_column = None
        self.url_column = None
        self.group_column = None
        self.subgroup_column = None
        self.surname_column = None
        self.firstname_column = None

        for col in range(1, empty_column):

            #if self.sheet.range((1,col)).value == "username":
            if self.sheet.cell(row = 1,column = col).value  == "username":
                self.username_column = col

            #if self.sheet.range((1,col)).value == "password":
            if self.sheet.cell(row = 1,column = col).value  == "password":
                self.password_column = col

            #if self.sheet.range((1,col)).value == "repository_url":
            if self.sheet.cell(row = 1,column = col).value == "repository_url":
                self.url_column = col

            #if self.sheet.range((1,col)).value == "group":
            if self.sheet.cell(row = 1,column = col).value == "group":
                self.group_column = col

            #if self.sheet.range((1,col)).value == "subgroup":
            if self.sheet.cell(row = 1,column = col).value == "subgroup":
                self.subgroup_column = col

            #if self.sheet.range((1,col)).value.casefold() == "cognome":
            if self.sheet.cell(row = 1,column = col).value == "cognome":
                self.surname_column = col

            #if self.sheet.range((1,col)).value.casefold() == "nome":
            if self.sheet.cell(row = 1,column = col).value == "nome":
                self.firstname_column = col

        #self.num_students = self.sheet.range('A1').end('down').row - 1
        self.xlsx_num_students = len(list(self.sheet.rows)) - 1


        if self.surname_column is None or self.firstname_column is None:
            raise ValueError("Error: Missing name or surname columns in XLSX")


        if self.username_column is None:

            self.username_column = empty_column
            self.password_column = empty_column + 1
            self.group_column = empty_column + 2
            self.subgroup_column = empty_column + 3
            self.url_column = empty_column + 4


        count_students = "SELECT MAX(id) FROM students"
        max_students = self.connection.execute(count_students)
        student_id = max_students.fetchone()+1 if max_students.rowcount > 0 else 1

        student_id = initial_user_id if initial_user_id > student_id else student_id


        cursor = self.connection.cursor()

        for row in range(2, self.xlsx_num_students+2):

            #username = self.sheet.range((row, self.username_column)).value
            #password = self.sheet.range((row, self.password_column)).value
            username = self.sheet.cell(row = row, column = self.username_column).value
            password = self.sheet.cell(row = row, column = self.password_column).value

            if username is None and password is None:

                username = prefix_username + str(student_id)

                password = ''.join(secrets.choice(alphabet) for i in range(password_length))


            print(f"Initializing user '{username}'")

            insert_student = "INSERT INTO students VALUES (?,?,?,?,?,'','','','',0)"

            cursor.execute(insert_student, (student_id, username, password, top_project_group, project_subgroup))

            student_id += 1

        self.connection.commit()
        cursor.close()

        self.wb.close()

    def delete_user(self, username: str):

        cursor = self.connection.cursor()

        delete_student = "DELETE FROM students WHERE username=?"

        cursor.execute(delete_student, (username,))

        self.connection.commit()
        cursor.close()




class StudentsIter:

    def __init__(self, students: StudentsDB):

        self.students = students
        self._current_index = 0
        self._num_students = students.get_num_students()
        self._results = students.connection.execute("SELECT username, password, firstname, surname, repository_url FROM students")


    def __iter__(self):
        return self

    def __next__(self):

        if self._current_index < self._num_students:

            row = self._results.fetchone()

            student = {
                        #"row": row,
                        "username": row[0],
                        "password": row[1],
                        "firstname": row[2],
                        "surname": row[3],
                        "repository_url": row[4]
                    }

            self._current_index += 1
            return student

        raise StopIteration

