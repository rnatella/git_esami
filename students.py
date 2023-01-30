#import xlwings as xw
from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Font
# using this because xlwings does not work in ubuntu
import os
import secrets
import string

class Students:

    def __init__(self, xlsx_path: str) -> None:

        if not os.path.exists(xlsx_path):
            raise ValueError(f"Error: '{xlsx_path}' does not exist")
            

        if not xlsx_path.endswith('.xlsx'):
            raise ValueError(f"Error: '{xlsx_path}' does not seem an XLSX file")

        #self.wb = xw.Book(xlsx_path)
        #self.sheet = self.wb.sheets[0]
        self.wb = load_workbook(xlsx_path)
        self.sheet = self.wb.worksheets[0]
        self.xlsx_path = xlsx_path

        #empty_column = self.sheet.used_range[-1].offset(column_offset=1).column
        empty_column = len(list(self.sheet.columns)) + 1

        self.username_column = None
        self.password_column = None
        self.url_column = None
        self.group_column = None
        self.subgroup_column = None
        self.surname_column = None
        self.firstname_column = None

        for col in range(1, empty_column):

            #if self.sheet.range((1,col)).value == "username":
            if self.sheet.cell(row = 1,column = col).value  == "username":
                self.username_column = col

            #if self.sheet.range((1,col)).value == "password":
            if self.sheet.cell(row = 1,column = col).value  == "password":
                self.password_column = col

            #if self.sheet.range((1,col)).value == "repository_url":
            if self.sheet.cell(row = 1,column = col).value == "repository_url":
                self.url_column = col

            #if self.sheet.range((1,col)).value == "group":
            if self.sheet.cell(row = 1,column = col).value == "group":
                self.group_column = col

            #if self.sheet.range((1,col)).value == "subgroup":
            if self.sheet.cell(row = 1,column = col).value == "subgroup":
                self.subgroup_column = col

            #if self.sheet.range((1,col)).value.casefold() == "cognome":
            if self.sheet.cell(row = 1,column = col).value == "cognome":
                self.surname_column = col

            #if self.sheet.range((1,col)).value.casefold() == "nome":
            if self.sheet.cell(row = 1,column = col).value == "nome":
                self.firstname_column = col

        #self.num_students = self.sheet.range('A1').end('down').row - 1
        self.num_students = len(list(self.sheet.rows)) - 1


        if self.surname_column is None or self.firstname_column is None:
            raise ValueError("Error: Missing name or surname columns in XLSX")


        if self.username_column is None:

            #print("Inizializing XLSX columns...")
            
            cell_font = Font( bold = True )

            self.username_column = empty_column
            self.password_column = empty_column + 1
            self.group_column = empty_column + 2
            self.subgroup_column = empty_column + 3
            self.url_column = empty_column + 4

            #self.sheet.range((1, self.username_column)).value = "username"
            #self.sheet.range((1, self.username_column)).font.bold = True
            self.sheet.cell(row = 1, column = self.username_column).value = "username"
            self.sheet.cell(row = 1, column = self.username_column).font = cell_font

            #self.sheet.range((1, self.password_column)).value = "password"
            #self.sheet.range((1, self.password_column)).font.bold = True
            self.sheet.cell(row = 1, column = self.password_column).value = "password"
            self.sheet.cell(row = 1, column = self.password_column).font = cell_font

            #self.sheet.range((1, self.group_column)).value = "group"
            #self.sheet.range((1, self.group_column)).font.bold = True
            self.sheet.cell(row = 1, column = self.password_column).value = "group"
            self.sheet.cell(row = 1, column = self.password_column).font = cell_font

            #self.sheet.range((1, self.subgroup_column)).value = "subgroup"
            #self.sheet.range((1, self.subgroup_column)).font.bold = True
            self.sheet.cell(row = 1, column = self.subgroup_column).value = "sbugroup"
            self.sheet.cell(row = 1, column = self.subgroup_column).font = cell_font

            #self.sheet.range((1, self.url_column)).value = "repository_url"
            #self.sheet.range((1, self.url_column)).font.bold = True
            self.sheet.cell(row = 1, column = self.url_column).value = "repository_url"
            self.sheet.cell(row = 1, column = self.url_column).font = cell_font


    def get_num_students(self) -> int:
        return self.num_students

    def set_repository_url(self, row: int, repository_url: str) -> None:
        #self.sheet.range((row, self.url_column)).value = repository_url
        self.sheet.cell(row = row, column = self.url_column).value = repository_url


    def initialize_users(self, prefix_username: str, initial_user_id: int, password_length: int, top_project_group: str, project_subgroup: str) -> None:

        alphabet = string.ascii_letters + string.digits

        student_id = initial_user_id

        for row in range(2, self.num_students+2):

            #username = self.sheet.range((row, self.username_column)).value
            #password = self.sheet.range((row, self.password_column)).value
            username = self.sheet.cell(row = row, column = self.username_column).value
            password = self.sheet.cell(row = row, column = self.password_column).value


            if username is None and password is None:

                username = prefix_username + str(initial_user_id)
                initial_user_id += 1

                password = ''.join(secrets.choice(alphabet) for i in range(password_length))

                print(f"Initializing user '{username}'")

                #self.sheet.range((row, self.username_column)).value = username
                #self.sheet.range((row, self.password_column)).value = password
                self.sheet.cell(row = row, column = self.username_column).value = username
                self.sheet.cell(row = row, column = self.password_column).value = password

                #self.sheet.range((row, self.group_column)).value = top_project_group
                #self.sheet.range((row, self.subgroup_column)).value = project_subgroup
                self.sheet.cell(row = row, column = self.group_column).value = top_project_group
                self.sheet.cell(row = row, column = self.subgroup_column).value = project_subgroup


        self.wb.save(self.xlsx_path)        
       


    def __del__(self):
        self.wb.save(self.xlsx_path)
        self.wb.close()

    def __iter__(self):
        return StudentsIter(self)



class StudentsIter:

    def __init__(self, students: Students):
        self._num_students = students.num_students
        self._sheet = students.sheet

        self._username_column = students.username_column
        self._password_column = students.password_column
        self._firstname_column = students.firstname_column
        self._surname_column = students.surname_column
        self._url_column = students.url_column

        self._current_index = 0

    def __iter__(self):
        return self

    def __next__(self):
        if self._current_index < self._num_students:

            row = self._current_index + 2

            #username  = self._sheet.range((row, self._username_column)).value
            #password  = self._sheet.range((row, self._password_column)).value
            
            username  = self._sheet.cell(row = row, column = self._username_column).value
            password  = self._sheet.cell(row = row, column = self._password_column).value

            #firstname = self._sheet.range((row, self._firstname_column)).value
            #surname   = self._sheet.range((row, self._surname_column)).value
            
            firstname = self._sheet.cell(row = row, column = self._firstname_column).value
            surname   = self._sheet.cell(row = row, column = self._surname_column).value

            #repository_url = self._sheet.range((row, self._url_column)).value
            repository_url = self._sheet.cell(row = row, column = self._url_column).value

            student = {
                        "row": row,
                        "username": username, 
                        "password": password, 
                        "firstname": firstname, 
                        "surname": surname,
                        "repository_url": repository_url
                    }

            self._current_index += 1
            return student
        
        raise StopIteration
